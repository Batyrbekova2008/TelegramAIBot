"""
Benchmark service: Groq vs Ollama (Task 15)

/benchmark command runs 20 standard questions on both platforms,
collects latency (p50, p95, p99), tokens/sec, and exports Excel report.
"""

import asyncio
import json
import logging
import statistics
import time
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from groq import AsyncGroq
    from services.ollama_service import OllamaService

log = logging.getLogger("benchmark")

# fmt: off
BENCHMARK_QUESTIONS = [
    "Что такое рекурсия?",
    "Объясни разницу между стеком и очередью.",
    "Как работает бинарный поиск?",
    "Что такое Big O нотация?",
    "Объясни принципы ООП.",
    "Что такое HTTP и HTTPS?",
    "Как работает TCP/IP?",
    "Что такое REST API?",
    "Объясни транзакции в базах данных.",
    "Что такое индекс в SQL?",
    "Как работает Git?",
    "Что такое Docker?",
    "Объясни разницу между процессом и потоком.",
    "Что такое виртуальная память?",
    "Как работает сборщик мусора?",
    "Что такое deadlock?",
    "Объясни алгоритм сортировки слиянием.",
    "Что такое хеш-таблица?",
    "Как работает публичный ключ шифрования?",
    "Что такое микросервисная архитектура?",
]


async def _time_request(coro) -> tuple[float, str, int]:
    """Returns (latency_seconds, text_response, token_count)."""
    start = time.perf_counter()
    try:
        text = await coro
        elapsed = time.perf_counter() - start
        tokens = len(text.split())
        return elapsed, text, tokens
    except Exception as e:
        elapsed = time.perf_counter() - start
        return elapsed, f"ERROR: {e}", 0


async def _percentile(values: list[float], p: float) -> float:
    if not values:
        return 0.0
    sorted_vals = sorted(values)
    idx = (len(sorted_vals) - 1) * p / 100
    lo, hi = int(idx), min(int(idx) + 1, len(sorted_vals) - 1)
    return sorted_vals[lo] + (sorted_vals[hi] - sorted_vals[lo]) * (idx - lo)


class BenchmarkService:
    def __init__(self, groq_client: "AsyncGroq", ollama: "OllamaService"):
        self.groq = groq_client
        self.ollama = ollama

    async def run(
        self,
        questions: list[str] | None = None,
        groq_model: str = "llama-3.1-8b-instant",
        ollama_model: str = "llama3.2:3b",
        progress_callback=None,
    ) -> dict:
        questions = questions or BENCHMARK_QUESTIONS
        total = len(questions)
        groq_results = []
        ollama_results = []
        ollama_available = await self.ollama.is_available()

        for i, question in enumerate(questions, 1):
            if progress_callback:
                await progress_callback(i, total, question)

            # Groq
            async def _ask_groq(q=question):
                r = await self.groq.chat.completions.create(
                    model=groq_model,
                    messages=[{"role": "user", "content": q}],
                    timeout=15,
                )
                return r.choices[0].message.content or ""

            lat, text, tok = await _time_request(_ask_groq())
            groq_results.append({
                "question": question,
                "latency": lat,
                "tokens": tok,
                "tps": tok / lat if lat > 0 else 0,
                "answer": text[:200],
                "error": text.startswith("ERROR:"),
            })

            # Ollama
            if ollama_available:
                async def _ask_ollama(q=question):
                    return await self.ollama.chat(
                        messages=[{"role": "user", "content": q}],
                        model=ollama_model,
                    )
                lat2, text2, tok2 = await _time_request(_ask_ollama())
                ollama_results.append({
                    "question": question,
                    "latency": lat2,
                    "tokens": tok2,
                    "tps": tok2 / lat2 if lat2 > 0 else 0,
                    "answer": text2[:200],
                    "error": text2.startswith("ERROR:"),
                })

            await asyncio.sleep(0.2)  # avoid rate limits

        def _stats(results):
            lats = [r["latency"] for r in results if not r["error"]]
            tpss = [r["tps"] for r in results if not r["error"]]
            return {
                "count": len(results),
                "errors": sum(1 for r in results if r["error"]),
                "p50_latency": asyncio.get_event_loop().run_until_complete(
                    asyncio.coroutine(lambda: _percentile(lats, 50))()
                ) if lats else 0,
                "p95_latency": 0,
                "p99_latency": 0,
                "mean_tps": statistics.mean(tpss) if tpss else 0,
            }

        # Compute stats synchronously
        def sync_stats(results):
            lats = [r["latency"] for r in results if not r["error"]]
            tpss = [r["tps"] for r in results if not r["error"]]
            if not lats:
                return {"count": len(results), "errors": len(results),
                        "p50_latency": 0, "p95_latency": 0, "p99_latency": 0, "mean_tps": 0}
            lats_sorted = sorted(lats)
            n = len(lats_sorted)
            p = lambda pct: lats_sorted[min(int((n - 1) * pct / 100), n - 1)]
            return {
                "count": len(results),
                "errors": sum(1 for r in results if r["error"]),
                "p50_latency": round(p(50), 3),
                "p95_latency": round(p(95), 3),
                "p99_latency": round(p(99), 3),
                "mean_tps": round(statistics.mean(tpss), 1) if tpss else 0,
            }

        return {
            "timestamp": datetime.utcnow().isoformat(),
            "groq_model": groq_model,
            "ollama_model": ollama_model if ollama_available else "N/A (unavailable)",
            "groq": {"stats": sync_stats(groq_results), "results": groq_results},
            "ollama": {"stats": sync_stats(ollama_results), "results": ollama_results}
            if ollama_available else {"stats": {}, "results": []},
        }

    def to_excel(self, benchmark_data: dict) -> bytes:
        """Export benchmark results to Excel. Returns bytes."""
        wb = openpyxl.Workbook()

        # ── Summary sheet ──────────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Summary"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2B4590")

        headers = ["Metric", "Groq", "Ollama"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = 22

        groq_s = benchmark_data["groq"]["stats"]
        ollama_s = benchmark_data["ollama"]["stats"]
        rows = [
            ("Model", benchmark_data["groq_model"], benchmark_data["ollama_model"]),
            ("Total questions", groq_s.get("count", 0), ollama_s.get("count", 0)),
            ("Errors", groq_s.get("errors", 0), ollama_s.get("errors", 0)),
            ("p50 latency (s)", groq_s.get("p50_latency", 0), ollama_s.get("p50_latency", 0)),
            ("p95 latency (s)", groq_s.get("p95_latency", 0), ollama_s.get("p95_latency", 0)),
            ("p99 latency (s)", groq_s.get("p99_latency", 0), ollama_s.get("p99_latency", 0)),
            ("Avg tokens/sec", groq_s.get("mean_tps", 0), ollama_s.get("mean_tps", 0)),
        ]
        for r, row in enumerate(rows, 2):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)

        # ── Groq detail sheet ──────────────────────────────────────────────────
        self._add_detail_sheet(wb, "Groq Results", benchmark_data["groq"]["results"])
        self._add_detail_sheet(wb, "Ollama Results", benchmark_data["ollama"]["results"])

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def _add_detail_sheet(wb, title: str, results: list[dict]):
        ws = wb.create_sheet(title=title)
        header_font = Font(bold=True)
        cols = ["#", "Question", "Latency (s)", "Tokens", "Tokens/sec", "Error", "Answer"]
        for c, h in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["G"].width = 50

        for r, item in enumerate(results, 2):
            ws.cell(row=r, column=1, value=r - 1)
            ws.cell(row=r, column=2, value=item.get("question", ""))
            ws.cell(row=r, column=3, value=item.get("latency", 0))
            ws.cell(row=r, column=4, value=item.get("tokens", 0))
            ws.cell(row=r, column=5, value=round(item.get("tps", 0), 1))
            ws.cell(row=r, column=6, value="Yes" if item.get("error") else "No")
            ws.cell(row=r, column=7, value=item.get("answer", "")[:200])
